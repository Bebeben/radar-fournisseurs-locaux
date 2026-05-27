#!/usr/bin/env python3
"""
Outil de recherche de producteurs locaux pour Super U.

Recherche les producteurs, artisans et fournisseurs locaux autour d'une ville
donnée et génère un fichier Excel de prospection.

Usage :
    python main.py
    python main.py --city "Cherbourg, Manche, France" --radius 30
    python main.py --google-api-key YOUR_KEY
"""

import argparse
import logging
import os
import sys
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import config
from scrapers.bienvenue_ferme import BienvenueFermeScraper
from scrapers.google_places import GooglePlacesScraper
from scrapers.pages_jaunes import PagesJaunesScraper
from scrapers.acheteralasource import AcheterALaSourceScraper
from scrapers.locavor import LocavorScraper
from enrichment.competitor_checker import check_competitors
from enrichment.contact_enricher import enrich_contacts
from utils.geocoding import geocode_records
from utils.categorizer import categorize_records
from utils.dedup import deduplicate

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
#  Excel writer
# ──────────────────────────────────────────────

EXCEL_FIELD_MAP = {
    "Nom du producteur": "nom",
    "Raison sociale": "raison_sociale",
    "Catégorie": "categorie",
    "Sous-catégorie": "sous_categorie",
    "Produits phares": "produits",
    "Ville": "ville",
    "Code postal": "code_postal",
    "Distance (km)": "distance_km",
    "Téléphone": "telephone",
    "Email": "email",
    "Site web": "site_web",
    "Réseaux sociaux": "reseaux_sociaux",
    "Labels / Certifications": "labels",
    "Vu chez Leclerc Equeurdreville": "vu_chez_leclerc_equeurdreville",
    "Vu chez Intermarché Les Pieux": "vu_chez_intermarche_les_pieux",
    "Source": "source",
    "Date de collecte": "date_collecte",
}


def load_existing_excel(path: str) -> list[dict]:
    """Charge un fichier Excel existant en liste de dicts."""
    if not os.path.exists(path):
        return []
    try:
        df = pd.read_excel(path, engine="openpyxl")
        # Mapper les noms de colonnes Excel vers les noms internes
        reverse_map = {v_excel: v_key for v_excel, v_key in EXCEL_FIELD_MAP.items()}
        df = df.rename(columns=reverse_map)
        # Remplacer NaN par des chaînes vides pour les champs texte
        text_cols = [c for c in df.columns if c not in ("latitude", "longitude", "distance_km")]
        for col in text_cols:
            df[col] = df[col].fillna("")
        records = df.to_dict(orient="records")
        logger.info(f"Fichier existant chargé : {len(records)} enregistrements")
        return records
    except Exception as e:
        logger.warning(f"Impossible de lire le fichier existant : {e}")
        return []


def write_excel(records: list[dict], output_path: str):
    """Écrit les enregistrements dans un fichier Excel formaté."""
    if not records:
        logger.warning("Aucun enregistrement à écrire")
        return

    # Préparer les données pour le DataFrame
    rows = []
    for rec in records:
        row = {}
        for col_name, field in EXCEL_FIELD_MAP.items():
            val = rec.get(field, "")
            if val is None:
                val = ""
            # Code postal : forcer en string sans décimales
            if field == "code_postal" and val:
                try:
                    val = str(int(float(val)))
                except (ValueError, TypeError):
                    val = str(val)
            # Distance : arrondir
            if field == "distance_km" and val:
                try:
                    val = round(float(val), 1)
                except (ValueError, TypeError):
                    pass
            row[col_name] = val
        rows.append(row)

    df = pd.DataFrame(rows, columns=list(EXCEL_FIELD_MAP.keys()))

    # Trier par distance
    df["_dist_sort"] = pd.to_numeric(df["Distance (km)"], errors="coerce")
    df = df.sort_values("_dist_sort", na_position="last").drop(columns=["_dist_sort"])
    df = df.reset_index(drop=True)

    # Écrire avec pandas puis formater avec openpyxl
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Formatage openpyxl
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Producteurs Locaux"

    # --- Styles ---
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # En-têtes
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Bordures et alignement sur toutes les cellules
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze panes (figer la ligne d'en-tête)
    ws.freeze_panes = "A2"

    # Auto-filtre
    ws.auto_filter.ref = ws.dimensions

    # Largeur des colonnes
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        width = min(max_length + 2, 45)
        width = max(width, 10)
        ws.column_dimensions[col_letter].width = width

    # Mise en forme conditionnelle — colonnes concurrentielles
    col_names = list(EXCEL_FIELD_MAP.keys())

    for col_name in ["Vu chez Leclerc Equeurdreville", "Vu chez Intermarché Les Pieux"]:
        if col_name in col_names:
            col_idx = col_names.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            range_str = f"{col_letter}2:{col_letter}{ws.max_row}"

            # Vert si "Non" (pas chez le concurrent = opportunité)
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(
                    operator="equal",
                    formula=['"Non"'],
                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    font=Font(color="006100"),
                ),
            )
            # Rouge si "Oui" (déjà chez le concurrent)
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(
                    operator="equal",
                    formula=['"Oui"'],
                    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    font=Font(color="9C0006"),
                ),
            )

    # Hyperliens sur la colonne Site web
    if "Site web" in col_names:
        web_col = col_names.index("Site web") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=web_col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")

    wb.save(output_path)
    logger.info(f"Fichier Excel écrit : {output_path} ({len(records)} lignes)")


# ──────────────────────────────────────────────
#  Pipeline principal
# ──────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Recherche de producteurs locaux pour Super U"
    )
    parser.add_argument(
        "--city", default=None,
        help=f"Ville de référence (défaut: {config.TARGET_CITY})"
    )
    parser.add_argument(
        "--radius", type=int, default=None,
        help=f"Rayon de recherche en km (défaut: {config.RADIUS_KM})"
    )
    parser.add_argument(
        "--google-api-key", default=None,
        help="Clé API Google Places (alternative à la variable d'environnement)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Chemin du fichier Excel de sortie"
    )
    parser.add_argument(
        "--sources", default=None,
        help="Sources à utiliser, séparées par des virgules (ex: bienvenue,google,locavor)"
    )
    return parser.parse_args()


def main():
    # Logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    args = parse_args()

    # Appliquer les arguments CLI
    if args.city:
        config.TARGET_CITY = args.city
        logger.info(f"Ville de référence : {config.TARGET_CITY}")
        # Re-géocoder la ville cible
        from utils.geocoding import geocode
        coords = geocode(args.city)
        if coords:
            config.TARGET_LAT, config.TARGET_LNG = coords
            logger.info(f"Coordonnées : {config.TARGET_LAT}, {config.TARGET_LNG}")
        else:
            logger.error(f"Impossible de géocoder '{args.city}'. Utilisation des coordonnées par défaut.")

    if args.radius:
        config.RADIUS_KM = args.radius

    if args.google_api_key:
        config.GOOGLE_API_KEY = args.google_api_key

    output_path = args.output or os.path.join(config.OUTPUT_DIR, config.OUTPUT_FILE)

    # Sélection des sources
    all_scrapers = {
        "bienvenue": BienvenueFermeScraper,
        "google": GooglePlacesScraper,
        "pagesjaunes": PagesJaunesScraper,
        "acheteralasource": AcheterALaSourceScraper,
        "locavor": LocavorScraper,
    }

    if args.sources:
        selected = [s.strip().lower() for s in args.sources.split(",")]
        scrapers = [cls() for name, cls in all_scrapers.items() if name in selected]
    else:
        scrapers = [cls() for cls in all_scrapers.values()]

    # ── Étape 1 : Scraping ──
    logger.info("=" * 60)
    logger.info(f"RECHERCHE DE PRODUCTEURS LOCAUX — {config.TARGET_CITY}")
    logger.info(f"Rayon : {config.RADIUS_KM} km | Sources : {len(scrapers)}")
    logger.info("=" * 60)

    all_records = []
    for scraper in scrapers:
        try:
            records = scraper.scrape()
            logger.info(f"✓ {scraper.source_name} : {len(records)} producteurs trouvés")
            all_records.extend(records)
        except Exception as e:
            logger.error(f"✗ {scraper.source_name} : échec — {e}")

    logger.info(f"\nTotal brut : {len(all_records)} enregistrements")

    if not all_records:
        logger.warning("Aucun résultat trouvé. Vérifiez votre connexion et les paramètres.")
        sys.exit(1)

    # ── Étape 2 : Géocodage ──
    logger.info("\n--- Géocodage et calcul des distances ---")
    geocode_records(all_records)

    # ── Étape 3 : Catégorisation ──
    logger.info("\n--- Catégorisation ---")
    categorize_records(all_records)

    # ── Étape 4 : Chargement existant + Dédoublonnage ──
    logger.info("\n--- Dédoublonnage ---")
    existing = load_existing_excel(output_path)
    merged = deduplicate(all_records, existing)
    logger.info(f"Après dédoublonnage : {len(merged)} producteurs uniques")

    # ── Étape 5 : Enrichissement contacts (web) ──
    logger.info("\n--- Enrichissement contacts (site web / recherche Google) ---")
    enrich_contacts(merged)

    # ── Étape 6 : Enrichissement concurrentiel ──
    logger.info("\n--- Vérification concurrentielle ---")
    check_competitors(merged)

    # ── Étape 7 : Export Excel ──
    logger.info("\n--- Génération du fichier Excel ---")
    write_excel(merged, output_path)

    # Résumé
    logger.info("\n" + "=" * 60)
    logger.info("TERMINÉ")
    logger.info(f"  Producteurs : {len(merged)}")
    logger.info(f"  Fichier     : {os.path.abspath(output_path)}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
