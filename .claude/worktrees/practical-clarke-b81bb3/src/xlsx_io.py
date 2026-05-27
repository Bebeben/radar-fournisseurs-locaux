"""
xlsx_io
=======

Lecture / écriture du fichier Excel local `commande_prix_station_v2.xlsx`.

Utilisé tant que Benjamin n'a pas le Google Sheet (avant prise de fonction).
Après cession et conversion en Google Sheet natif, ce module sera remplacé
par l'usage de Google Sheets API (cf. sheet_io.py).

Onglets gérés ici :
- **Stations**     : lecture seule (Benjamin l'édite à la main)
- **Concurrents**  : append-only à chaque run, 1 ligne par station relevée
- **Reco prix**    : append-only à chaque run, 1 ligne avec recommandation
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.moteur_decision import PrixCarburants, Proposition

log = logging.getLogger(__name__)

# Headers de l'onglet "Stations" (ordre des colonnes A à G)
COL_STATIONS = ["Type", "Nom", "ID prix-carburants", "Code postal", "Distance km", "Alignement actif", "Notes"]

# Headers "Concurrents" (A à K)
COL_CONCURRENTS = ["Date", "Heure", "Station", "Type", "Distance km",
                   "E85", "SP95-E10", "SP98", "Gazole", "E85", "Notes"]

# Headers "Reco prix" (A à L)
COL_RECO_PRIX = ["Date", "Heure", "Statut",
                 "E85 TTC", "SP95-E10 TTC", "SP98 TTC", "Gazole TTC",
                 "Marge SP95 %", "Marge SP95-E10 %", "Marge SP98 %", "Marge Gazole %",
                 "Raison"]


class XlsxIO:
    """Wrapper openpyxl pour le fichier Excel local."""

    def __init__(self, xlsx_path: str | Path):
        self.xlsx_path = Path(xlsx_path)
        if not self.xlsx_path.is_file():
            raise FileNotFoundError(f"Fichier Excel introuvable : {self.xlsx_path}")

    # ------------------------------------------------------------
    # Lecture
    # ------------------------------------------------------------

    def lire_stations(self) -> list[dict[str, Any]]:
        """Lit l'onglet Stations et renvoie une liste de dicts.

        Chaque dict a les clés : type, nom, id_prix_carburants, code_postal,
        distance_km, alignement_actif (bool), notes.
        """
        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        if "Stations" not in wb.sheetnames:
            raise ValueError("Onglet 'Stations' introuvable dans le xlsx")
        ws = wb["Stations"]
        stations: list[dict[str, Any]] = []
        # Skip ligne 1 (headers)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                continue  # ligne vide
            stations.append({
                "type": str(row[0]).strip(),
                "nom": str(row[1]).strip() if row[1] else "",
                "id_prix_carburants": str(row[2]).strip() if row[2] else "",
                "code_postal": str(row[3]).strip() if row[3] else "",
                "distance_km": float(row[4]) if row[4] is not None else 0.0,
                "alignement_actif": _parse_bool(row[5]),
                "notes": str(row[6]) if row[6] else "",
            })
        return stations

    # ------------------------------------------------------------
    # Écriture
    # ------------------------------------------------------------

    def append_concurrents(
        self,
        prix_par_station: dict[str, PrixCarburants],
        stations_meta: list[dict[str, Any]],
        prix_e85_par_station: dict[str, float | None],
        derniere_maj_par_station: dict[str, str | None] | None = None,
        maintenant: datetime | None = None,
    ) -> int:
        """Append 1 ligne par station relevée dans l'onglet Concurrents.

        Args:
            prix_par_station: id_prix_carburants -> PrixCarburants
            stations_meta: liste des stations (issue de lire_stations()) — pour
                récupérer nom/type/distance.
            prix_e85_par_station: id_prix_carburants -> prix E85 (€/L) ou None.
                E85 n'est pas dans PrixCarburants pour l'instant.
            maintenant: timestamp du run (Europe/Paris).

        Returns:
            Nombre de lignes ajoutées.
        """
        wb = load_workbook(self.xlsx_path)
        if "Concurrents" not in wb.sheetnames:
            raise ValueError("Onglet 'Concurrents' introuvable")
        ws = wb["Concurrents"]

        # Trouver la prochaine ligne vide
        prochaine = ws.max_row + 1
        if ws.max_row == 1:  # juste les headers
            prochaine = 2

        nb_ajoutees = 0
        for meta in stations_meta:
            id_st = meta["id_prix_carburants"]
            if id_st not in prix_par_station:
                continue
            pc = prix_par_station[id_st]
            ws.cell(row=prochaine, column=1, value=maintenant.date().isoformat())
            ws.cell(row=prochaine, column=2, value=maintenant.strftime("%H:%M"))
            ws.cell(row=prochaine, column=3, value=meta["nom"])
            ws.cell(row=prochaine, column=4, value=meta["type"])
            ws.cell(row=prochaine, column=5, value=meta["distance_km"])
            ws.cell(row=prochaine, column=6, value=pc.e85)
            ws.cell(row=prochaine, column=7, value=pc.sp95_e10)
            ws.cell(row=prochaine, column=8, value=pc.sp98)
            ws.cell(row=prochaine, column=9, value=pc.gazole)
            ws.cell(row=prochaine, column=10, value=prix_e85_par_station.get(id_st))
            ws.cell(row=prochaine, column=11, value=meta.get("notes", ""))
            prochaine += 1
            nb_ajoutees += 1

        wb.save(self.xlsx_path)
        return nb_ajoutees

    def append_reco_prix(
        self,
        proposition: Proposition,
        nos_prix_ttc: PrixCarburants,
        prix_recommandes_ttc: dict[str, float],
        marge_ponderee: float | None = None,
        maintenant: datetime | None = None,
    ) -> None:
        """Append 1 ligne dans l'onglet Reco prix.

        Si proposition.action == ACTION : utilise prix_recommandes_ttc
        Sinon : utilise nos_prix_ttc (= statu quo)
        """
        from src.moteur_decision import Action

        wb = load_workbook(self.xlsx_path)
        if "Reco prix" not in wb.sheetnames:
            raise ValueError("Onglet 'Reco prix' introuvable")
        ws = wb["Reco prix"]

        prochaine = ws.max_row + 1
        if ws.max_row == 1:
            prochaine = 2

        # Prix à mettre = recommandés (si ACTION) sinon prix actuels
        if proposition.action == Action.ACTION and prix_recommandes_ttc:
            prix_finaux = prix_recommandes_ttc
        else:
            prix_finaux = {
                "E85": nos_prix_ttc.e85,
                "SP95-E10": nos_prix_ttc.sp95_e10,
                "SP98": nos_prix_ttc.sp98,
                "Gazole": nos_prix_ttc.gazole,
            }

        # Marges en % par carburant
        marges_pct = {
            "E85": _pct(proposition.marges.get("E85")),
            "SP95-E10": _pct(proposition.marges.get("SP95-E10")),
            "SP98": _pct(proposition.marges.get("SP98")),
            "Gazole": _pct(proposition.marges.get("Gazole")),
        }

        ws.cell(row=prochaine, column=1, value=maintenant.date().isoformat())
        ws.cell(row=prochaine, column=2, value=maintenant.strftime("%H:%M"))
        ws.cell(row=prochaine, column=3, value=proposition.action.value)
        ws.cell(row=prochaine, column=4, value=prix_finaux.get("E85"))
        ws.cell(row=prochaine, column=5, value=prix_finaux.get("SP95-E10"))
        ws.cell(row=prochaine, column=6, value=prix_finaux.get("SP98"))
        ws.cell(row=prochaine, column=7, value=prix_finaux.get("Gazole"))
        ws.cell(row=prochaine, column=8, value=marges_pct["E85"])
        ws.cell(row=prochaine, column=9, value=marges_pct["SP95-E10"])
        ws.cell(row=prochaine, column=10, value=marges_pct["SP98"])
        ws.cell(row=prochaine, column=11, value=marges_pct["Gazole"])
        ws.cell(row=prochaine, column=12, value=proposition.justification)

        # Mise en évidence visuelle si ACTION
        if proposition.action == Action.ACTION:
            from openpyxl.styles import Font, PatternFill
            for col in range(1, 13):
                c = ws.cell(row=prochaine, column=col)
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        # Format pourcentage sur les colonnes Marge
        for col in range(8, 12):
            c = ws.cell(row=prochaine, column=col)
            if c.value is not None:
                c.number_format = "0.00%"
        # Format prix sur les colonnes prix
        for col in range(4, 8):
            c = ws.cell(row=prochaine, column=col)
            if c.value is not None:
                c.number_format = "0.000"

        wb.save(self.xlsx_path)

    def lire_reco_prix_historique(self, nb_lignes: int = 50) -> list[dict]:
        """Lit les N dernières lignes de Reco prix (sert au verrou intra-journée)."""
        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        if "Reco prix" not in wb.sheetnames:
            return []
        ws = wb["Reco prix"]
        lignes: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            lignes.append({
                "Date": str(row[0]) if row[0] else "",
                "Heure run": str(row[1]) if row[1] else "",
                "Statut": str(row[2]) if row[2] else "",
            })
        return lignes[-nb_lignes:]


# ============================================================
# Helpers
# ============================================================


def _parse_bool(val: Any) -> bool:
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("OUI", "YES", "TRUE", "1", "VRAI", "X")


def _pct(marge: Any) -> float | None:
    """Renvoie marge_pourcent (déjà en ratio 0-1) ou None."""
    if marge is None:
        return None
    return getattr(marge, "marge_pourcent", None)
